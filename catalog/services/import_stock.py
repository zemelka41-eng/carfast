from __future__ import annotations

import hashlib
import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Any, BinaryIO, Iterable
from uuid import uuid4

import openpyxl
from django.db import transaction
from django.utils.text import slugify

from catalog.models import Category, City, Offer, Product, Series

IMPORT_LOGGER_NAME = "catalog.import_stock"


class StockRowError(ValueError):
    """Raised when a row is not importable."""


@dataclass(frozen=True)
class ParsedStockRow:
    row_number: int
    title: str
    brand_slug: str
    category_slug: str
    model_code: str
    config: str
    city_name: str
    city_slug: str
    qty: int
    price: Decimal | None
    vat: str
    year: int | None


@dataclass
class StockImportReport:
    file_name: str
    sheet_name: str
    batch_token: str

    parsed_rows: int = 0
    skipped_rows: int = 0

    created_series: int = 0
    created_categories: int = 0
    created_cities: int = 0
    created_products: int = 0
    updated_products: int = 0
    created_offers: int = 0
    updated_offers: int = 0
    deactivated_offers: int = 0

    errors: list[dict[str, Any]] | None = None

    def add_error(self, row_number: int | None, message: str) -> None:
        if self.errors is None:
            self.errors = []
        self.errors.append({"row": row_number, "message": message})


def import_stock(
    *,
    file: str | Path | BinaryIO,
    file_name: str | None = None,
    sheet: str | None = None,
    dry_run: bool = False,
    deactivate_missing: bool = False,
) -> StockImportReport:
    """Import offers/stocks from XLSX.

    Supports two formats:
    - "KARFAST" (Наименование/Модель/Комплектация/Цена с НДС, руб./Наличие)
    - normalized template (brand/category/title/model_code/config/city/qty/price/vat/year)

    The import is idempotent: offers are upserted by (product, city, price, year, vat)
    with qty aggregated by identical keys.
    """

    import_logger = _get_import_logger()

    resolved_name = file_name
    if not resolved_name:
        if isinstance(file, (str, Path)):
            resolved_name = Path(file).name
        else:
            resolved_name = "uploaded.xlsx"

    workbook = _load_workbook(file, import_logger)
    sheet_obj = _select_sheet(workbook, sheet, import_logger)

    batch_token = uuid4().hex
    report = StockImportReport(
        file_name=resolved_name,
        sheet_name=sheet_obj.title,
        batch_token=batch_token,
    )

    parsed_rows = _parse_rows(sheet_obj, report)
    report.parsed_rows = len(parsed_rows)

    # Aggregate qty by offer-key
    aggregated: dict[tuple[str, str, str, str, str, Decimal | None, int | None, str], ParsedStockRow] = {}
    qty_by_key: dict[tuple[str, str, str, str, str, Decimal | None, int | None, str], int] = {}

    for row in parsed_rows:
        key = (
            row.brand_slug,
            row.category_slug,
            row.model_code,
            row.config,
            row.city_slug,
            row.price,
            row.year,
            row.vat,
        )
        aggregated.setdefault(key, row)
        qty_by_key[key] = qty_by_key.get(key, 0) + int(row.qty)

    # Pre-compute per-product stats (to keep Product.price/availability in sync)
    product_stats: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for key, qty in qty_by_key.items():
        brand_slug, category_slug, model_code, config, _city_slug, price, _year, _vat = key
        pkey = (brand_slug, category_slug, model_code, config)
        stats = product_stats.setdefault(pkey, {"total_qty": 0, "min_price": None})
        stats["total_qty"] += int(qty)
        if price is not None:
            current_min = stats.get("min_price")
            stats["min_price"] = price if current_min is None else min(current_min, price)

    if dry_run:
        _dry_run_apply(aggregated, qty_by_key, product_stats, report)
        if deactivate_missing:
            report.deactivated_offers = _count_deactivate_missing(
                source_file=report.file_name,
                current_batch_token=batch_token,
            )
        return report

    with transaction.atomic():
        caches: _DbCaches = _DbCaches()

        for key, sample_row in aggregated.items():
            qty = qty_by_key[key]
            brand_slug, category_slug, model_code, config, city_slug, price, year, vat = key
            pkey = (brand_slug, category_slug, model_code, config)

            series = _get_or_create_series(brand_slug, caches, report)
            category = _get_or_create_category(category_slug, caches, report)
            city = _get_or_create_city(sample_row.city_name, city_slug, caches, report)

            stats = product_stats.get(pkey) or {"total_qty": 0, "min_price": None}
            product = _upsert_product(
                title=sample_row.title,
                series=series,
                category=category,
                model_code=model_code,
                config=config,
                total_qty=int(stats["total_qty"]),
                min_price=stats.get("min_price"),
                caches=caches,
                report=report,
            )

            _upsert_offer(
                product=product,
                city=city,
                qty=int(qty),
                price=price,
                year=year,
                vat=vat,
                source_file=report.file_name,
                batch_token=batch_token,
                source_row_hash=_hash_offer_key(
                    product_sku=product.sku,
                    city_slug=city.slug,
                    price=price,
                    year=year,
                    vat=vat,
                ),
                caches=caches,
                report=report,
            )

        if deactivate_missing:
            report.deactivated_offers = _deactivate_missing_offers(
                source_file=report.file_name,
                current_batch_token=batch_token,
            )

    return report


# -----------------
# Parsing utilities
# -----------------


def _parse_rows(sheet_obj, report: StockImportReport) -> list[ParsedStockRow]:
    header_cells = list(sheet_obj.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_cells[0]) if header_cells else []
    header_map = {
        _normalize_header(value): idx
        for idx, value in enumerate(headers)
        if _normalize_header(value)
    }

    is_normalized = _looks_like_normalized_template(header_map)
    if is_normalized:
        return _parse_rows_normalized(sheet_obj, header_map, report)

    is_karfast = _looks_like_karfast(header_map)
    if is_karfast:
        return _parse_rows_karfast(sheet_obj, header_map, report)

    # Fallback: fixed column letters (A,B,D,J,K) even if headers are missing.
    return _parse_rows_karfast(sheet_obj, header_map, report, fallback_positions=True)


def _looks_like_normalized_template(header_map: dict[str, int]) -> bool:
    return "model_code" in header_map and "city" in header_map


def _looks_like_karfast(header_map: dict[str, int]) -> bool:
    required = {"наименование", "модель", "комплектация", "наличие"}
    if not required.issubset(header_map):
        return False
    return any(key.startswith("цена") and "ндс" in key for key in header_map)


def _parse_rows_karfast(
    sheet_obj,
    header_map: dict[str, int],
    report: StockImportReport,
    *,
    fallback_positions: bool = False,
) -> list[ParsedStockRow]:
    # Default positions per real file: A,B,D,J,K
    idx_title = header_map.get("наименование", 0)
    idx_model = header_map.get("модель", 1)
    idx_config = header_map.get("комплектация", 3)

    idx_price = None
    for key, idx in header_map.items():
        if key.startswith("цена") and "ндс" in key:
            idx_price = idx
            break
    if idx_price is None:
        idx_price = 9

    idx_city = header_map.get("наличие", 10)

    if fallback_positions:
        idx_title, idx_model, idx_config, idx_price, idx_city = 0, 1, 3, 9, 10

    current_brand_slug: str | None = None
    current_category_slug: str | None = None

    parsed: list[ParsedStockRow] = []

    for row_number, row_values in enumerate(sheet_obj.iter_rows(min_row=2, values_only=True), start=2):
        title_raw = _normalize_spaces(_safe_get(row_values, idx_title))
        model_code_raw = _normalize_spaces(_safe_get(row_values, idx_model))
        config_raw = _normalize_spaces(_safe_get(row_values, idx_config))
        price_raw = _safe_get(row_values, idx_price)
        city_raw = _safe_get(row_values, idx_city)

        if _is_row_empty([title_raw, model_code_raw, config_raw, price_raw, city_raw]):
            report.skipped_rows += 1
            continue

        # Section header / divider: no model
        if not model_code_raw:
            current_brand_slug = _update_brand_state(title_raw, current_brand_slug)
            current_category_slug = _update_category_state(title_raw, current_category_slug)
            report.skipped_rows += 1
            continue

        try:
            model_code = _normalize_model_code(model_code_raw)
            config = _normalize_config(config_raw)
            year = _extract_year(config, model_code_raw)
            vat = "с НДС"
            city_name = _normalize_city_name(city_raw)
            city_slug = _normalize_city_slug(city_name)
            price = _parse_price(price_raw)

            brand_slug = _detect_brand_slug(title_raw, current_brand_slug)
            category_slug = _detect_category_slug(title_raw, current_category_slug)
            title = title_raw.strip() or f"{brand_slug.upper()} {model_code}"

            if not city_name:
                raise StockRowError("City is empty")

            parsed.append(
                ParsedStockRow(
                    row_number=row_number,
                    title=title,
                    brand_slug=brand_slug,
                    category_slug=category_slug,
                    model_code=model_code,
                    config=config,
                    city_name=city_name,
                    city_slug=city_slug,
                    qty=1,
                    price=price,
                    vat=vat,
                    year=year,
                )
            )
        except StockRowError as exc:
            report.add_error(row_number, str(exc))
        except Exception as exc:  # noqa: BLE001
            report.add_error(row_number, f"Unexpected error: {exc}")

    return parsed


def _parse_rows_normalized(
    sheet_obj,
    header_map: dict[str, int],
    report: StockImportReport,
) -> list[ParsedStockRow]:
    idx_brand = header_map.get("brand")
    idx_category = header_map.get("category")
    idx_title = header_map.get("title")
    idx_model = header_map.get("model_code")
    if idx_model is None:
        idx_model = header_map.get("model")
    idx_config = header_map.get("config")
    idx_city = header_map.get("city")
    idx_qty = header_map.get("qty")
    if idx_qty is None:
        idx_qty = header_map.get("quantity")
    idx_price = header_map.get("price")
    idx_vat = header_map.get("vat")
    idx_year = header_map.get("year")

    parsed: list[ParsedStockRow] = []

    for row_number, row_values in enumerate(sheet_obj.iter_rows(min_row=2, values_only=True), start=2):
        model_code_raw = _normalize_spaces(_safe_get(row_values, idx_model))
        if not model_code_raw and _is_row_empty(row_values):
            report.skipped_rows += 1
            continue

        # allow divider rows without model_code
        if not model_code_raw:
            report.skipped_rows += 1
            continue

        try:
            brand_raw = _normalize_spaces(_safe_get(row_values, idx_brand))
            category_raw = _normalize_spaces(_safe_get(row_values, idx_category))
            title_raw = _normalize_spaces(_safe_get(row_values, idx_title))
            config_raw = _normalize_spaces(_safe_get(row_values, idx_config))
            city_raw = _safe_get(row_values, idx_city)
            qty_raw = _safe_get(row_values, idx_qty)
            price_raw = _safe_get(row_values, idx_price)
            vat_raw = _normalize_spaces(_safe_get(row_values, idx_vat))
            year_raw = _safe_get(row_values, idx_year)

            model_code = _normalize_model_code(model_code_raw)
            config = _normalize_config(config_raw)

            brand_slug = _slugify_any(brand_raw) if brand_raw else "other"
            category_slug = _slugify_any(category_raw) if category_raw else "tehnika"

            city_name = _normalize_city_name(city_raw)
            if not city_name:
                raise StockRowError("City is empty")
            city_slug = _normalize_city_slug(city_name)

            qty = _parse_qty(qty_raw)
            price = _parse_price(price_raw)
            vat = vat_raw or "с НДС"

            year = _parse_year(year_raw)
            if year is None:
                year = _extract_year(config, model_code_raw)

            title = title_raw.strip() or f"{brand_slug.upper()} {model_code}"

            parsed.append(
                ParsedStockRow(
                    row_number=row_number,
                    title=title,
                    brand_slug=brand_slug or "other",
                    category_slug=category_slug or "tehnika",
                    model_code=model_code,
                    config=config,
                    city_name=city_name,
                    city_slug=city_slug,
                    qty=qty,
                    price=price,
                    vat=vat,
                    year=year,
                )
            )
        except StockRowError as exc:
            report.add_error(row_number, str(exc))
        except Exception as exc:  # noqa: BLE001
            report.add_error(row_number, f"Unexpected error: {exc}")

    return parsed


def _parse_qty(value: Any) -> int:
    if value is None or str(value).strip() == "":
        return 1
    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise StockRowError(f"Invalid qty '{value}'") from exc
    if parsed < 0:
        raise StockRowError(f"Invalid qty '{value}'")
    return parsed


def _parse_year(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        year = int(str(value).strip())
    except ValueError:
        return None
    if 1900 <= year <= 2100:
        return year
    return None


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^0-9a-zа-яё_\s]", "", text)
    text = text.replace(" ", "_")
    return text


def _normalize_spaces(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_model_code(value: str) -> str:
    model = (value or "").strip().upper()
    if not model:
        raise StockRowError("Model code is empty")
    return model


def _normalize_config(value: str) -> str:
    return _normalize_spaces(value)


def _normalize_city_name(value: Any) -> str:
    text = _normalize_spaces(value)
    if not text:
        return ""
    # strip leading "г." / "г"
    text = re.sub(r"^г\.?\s*", "", text, flags=re.IGNORECASE)
    return text.strip()


def _normalize_city_slug(city_name: str) -> str:
    slug = _slugify_any(city_name)
    if slug:
        return slug
    digest = hashlib.sha1(city_name.encode("utf-8")).hexdigest()[:10]
    return f"city-{digest}"


def _extract_year(*texts: str) -> int | None:
    for text in texts:
        if not text:
            continue
        match = re.search(r"(20\d{2})", str(text))
        if match:
            try:
                year = int(match.group(1))
            except ValueError:
                continue
            if 2000 <= year <= 2099:
                return year
    return None


def _parse_price(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except Exception:  # noqa: BLE001
            return None

    text = str(value).strip()
    if not text:
        return None

    # Keep digits and separators only.
    digits = re.sub(r"[^0-9.,]", "", text)
    if not digits:
        return None

    # For rubles we expect integer, but keep basic decimal support.
    if digits.count(",") == 1 and digits.count(".") == 0:
        digits = digits.replace(",", ".")
    else:
        digits = digits.replace(",", "").replace(".", "")

    try:
        return Decimal(digits).quantize(Decimal("0.01"))
    except Exception:  # noqa: BLE001
        return None


def _update_brand_state(title: str, current: str | None) -> str | None:
    text = (title or "").strip().upper()
    if not text:
        return current
    if "SHACMAN" in text:
        return "shacman"
    if "DAYUN" in text:
        return "dayun"
    return current


def _update_category_state(title: str, current: str | None) -> str | None:
    text = (title or "").strip().upper()
    if not text:
        return current
    if text.startswith("САМОСВАЛЫ"):
        return "samosvaly"
    if text.startswith("ТЯГАЧИ"):
        return "tyagachi"
    if text.startswith("АВТОБЕТОНОСМЕС"):
        return "abs"
    if text.startswith("КМУ"):
        return "kmu"
    if text.startswith("ЗЕРНОВОЗЫ"):
        return "zernovozy"
    return current


def _detect_brand_slug(title: str, current_brand_slug: str | None) -> str:
    text = (title or "").strip().upper()
    if text.startswith("DAYUN") or "DAYUN" in text:
        return "dayun"
    if "SHACMAN" in text:
        return "shacman"
    if current_brand_slug:
        return current_brand_slug
    return "other"


def _detect_category_slug(title: str, current_category_slug: str | None) -> str:
    text = (title or "").strip().lower()
    if text.startswith("самосвал"):
        return "samosvaly"
    if text.startswith("тягач"):
        return "tyagachi"
    if "автобетоносмес" in text:
        return "abs"
    if "кму" in text:
        return "kmu"
    if "зерновоз" in text:
        return "zernovozy"
    if "трактор" in text:
        return "traktory"
    if "фургон" in text:
        return "furgony"
    return current_category_slug or "tehnika"


def _is_row_empty(values: Iterable[Any]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return False
    return True


def _safe_get(row_values: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None:
        return None
    if idx < 0:
        return None
    if idx >= len(row_values):
        return None
    return row_values[idx]


# -----------------
# DB apply utilities
# -----------------


@dataclass
class _DbCaches:
    series_by_slug: dict[str, Series] = field(default_factory=dict)
    category_by_slug: dict[str, Category] = field(default_factory=dict)
    city_by_slug: dict[str, City] = field(default_factory=dict)
    product_by_sku_ci: dict[str, Product] = field(default_factory=dict)


_SERIES_NAMES: dict[str, str] = {
    "shacman": "SHACMAN",
    "dayun": "DAYUN",
    "other": "Other",
}

_CATEGORY_NAMES: dict[str, str] = {
    "samosvaly": "Самосвалы",
    "tyagachi": "Тягачи",
    "abs": "Автобетоносмесители",
    "kmu": "КМУ",
    "zernovozy": "Зерновозы",
    "traktory": "Тракторы",
    "furgony": "Фургоны",
    "tehnika": "Техника",
}


def _get_or_create_series(slug: str, caches: _DbCaches, report: StockImportReport) -> Series:
    slug = (slug or "other").strip().lower() or "other"
    if slug in caches.series_by_slug:
        return caches.series_by_slug[slug]

    obj = Series.objects.filter(slug__iexact=slug).first()
    if obj is None:
        obj = Series.objects.create(name=_SERIES_NAMES.get(slug, slug.upper()), slug=slug)
        report.created_series += 1
    caches.series_by_slug[slug] = obj
    return obj


def _get_or_create_category(slug: str, caches: _DbCaches, report: StockImportReport) -> Category:
    slug = (slug or "tehnika").strip().lower() or "tehnika"
    if slug in caches.category_by_slug:
        return caches.category_by_slug[slug]

    obj = Category.objects.filter(slug__iexact=slug).first()
    if obj is None:
        obj = Category.objects.create(name=_CATEGORY_NAMES.get(slug, slug), slug=slug)
        report.created_categories += 1
    caches.category_by_slug[slug] = obj
    return obj


def _get_or_create_city(
    name: str,
    slug: str,
    caches: _DbCaches,
    report: StockImportReport,
) -> City:
    slug = (slug or "").strip().lower() or _normalize_city_slug(name)
    if slug in caches.city_by_slug:
        return caches.city_by_slug[slug]

    obj = City.objects.filter(slug__iexact=slug).first()
    if obj is None:
        obj = City.objects.create(name=name or slug, slug=slug)
        report.created_cities += 1
    else:
        # Keep city name fresh (but don't override manual naming if identical).
        normalized_name = (name or "").strip()
        if normalized_name and obj.name != normalized_name:
            obj.name = normalized_name
            obj.save(update_fields=["name"])
    caches.city_by_slug[slug] = obj
    return obj


def _product_identity(
    *,
    brand_slug: str,
    category_slug: str,
    model_code: str,
    config: str,
) -> tuple[str, str]:
    key = f"{brand_slug}|{category_slug}|{model_code}|{config}".encode("utf-8")
    digest = hashlib.sha1(key).hexdigest()
    short = digest[:10].upper()

    model_part = re.sub(r"[^A-Z0-9]+", "-", model_code.upper()).strip("-") or "MODEL"
    model_part = model_part[:30]

    sku = f"{brand_slug.upper()}-{model_part}-{short}"
    sku = sku[:100]

    base = slugify(f"{brand_slug}-{model_code}") or slugify(f"{brand_slug}-{short.lower()}")
    base = (base or "product")[:40].strip("-")
    slug = f"{base}-{digest[:8]}"[:50].strip("-")

    return sku, slug


def _upsert_product(
    *,
    title: str,
    series: Series,
    category: Category,
    model_code: str,
    config: str,
    total_qty: int,
    min_price: Decimal | None,
    caches: _DbCaches,
    report: StockImportReport,
) -> Product:
    sku, computed_slug = _product_identity(
        brand_slug=series.slug,
        category_slug=category.slug,
        model_code=model_code,
        config=config,
    )
    sku_ci = sku.lower()
    if sku_ci in caches.product_by_sku_ci:
        product = caches.product_by_sku_ci[sku_ci]
    else:
        product = Product.objects.filter(sku__iexact=sku).first()
        if product:
            caches.product_by_sku_ci[sku_ci] = product

    if product is None:
        product = Product(
            sku=sku,
            slug=computed_slug,
            series=series,
            category=category,
            model_name_ru=title or sku,
            model_name_en=title or sku,
            model_code=model_code,
            config=config,
            price=min_price,
            availability=Product.Availability.IN_STOCK if total_qty > 0 else Product.Availability.ON_REQUEST,
            published=True,
            is_active=True,
        )
        product.save()
        report.created_products += 1
        caches.product_by_sku_ci[sku_ci] = product
        return product

    changed_fields: list[str] = []

    if product.series_id != series.id:
        product.series = series
        changed_fields.append("series")
    if product.category_id != category.id:
        product.category = category
        changed_fields.append("category")

    normalized_title = (title or "").strip()
    if normalized_title and product.model_name_ru != normalized_title:
        product.model_name_ru = normalized_title
        changed_fields.append("model_name_ru")
    if normalized_title and product.model_name_en != normalized_title:
        product.model_name_en = normalized_title
        changed_fields.append("model_name_en")

    if product.model_code != model_code:
        product.model_code = model_code
        changed_fields.append("model_code")
    if product.config != config:
        product.config = config
        changed_fields.append("config")

    desired_availability = (
        Product.Availability.IN_STOCK if total_qty > 0 else Product.Availability.ON_REQUEST
    )
    if product.availability != desired_availability:
        product.availability = desired_availability
        changed_fields.append("availability")

    if product.price != min_price:
        product.price = min_price
        changed_fields.append("price")

    if not product.slug:
        product.slug = computed_slug
        changed_fields.append("slug")

    if not product.is_active:
        product.is_active = True
        changed_fields.append("is_active")

    if changed_fields:
        product.save(update_fields=changed_fields)
        report.updated_products += 1

    caches.product_by_sku_ci[sku_ci] = product
    return product


def _upsert_offer(
    *,
    product: Product,
    city: City,
    qty: int,
    price: Decimal | None,
    year: int | None,
    vat: str,
    source_file: str,
    batch_token: str,
    source_row_hash: str,
    caches: _DbCaches,
    report: StockImportReport,
) -> Offer:
    vat = (vat or "с НДС").strip() or "с НДС"

    offer: Offer | None
    if price is None:
        offer = Offer.objects.filter(
            product=product,
            city=city,
            year=year,
            vat=vat,
            price__isnull=True,
        ).first()
        created = offer is None
        if offer is None:
            offer = Offer(
                product=product,
                city=city,
                price=None,
                year=year,
                vat=vat,
            )
    else:
        offer, created = Offer.objects.update_or_create(
            product=product,
            city=city,
            price=price,
            year=year,
            vat=vat,
            defaults={},
        )

    if created:
        report.created_offers += 1
    else:
        report.updated_offers += 1

    offer.qty = max(0, int(qty))
    offer.currency = "RUB"
    offer.source_file = source_file
    offer.source_row_hash = source_row_hash
    offer.batch_token = batch_token
    offer.is_active = True
    offer.save()
    return offer


def _hash_offer_key(
    *,
    product_sku: str,
    city_slug: str,
    price: Decimal | None,
    year: int | None,
    vat: str,
) -> str:
    payload = f"{product_sku}|{city_slug}|{price or ''}|{year or ''}|{vat}".encode("utf-8")
    return hashlib.sha1(payload).hexdigest()


def _deactivate_missing_offers(*, source_file: str, current_batch_token: str) -> int:
    return Offer.objects.filter(
        source_file=source_file,
        is_active=True,
    ).exclude(batch_token=current_batch_token).update(is_active=False)


def _count_deactivate_missing(*, source_file: str, current_batch_token: str) -> int:
    return (
        Offer.objects.filter(source_file=source_file, is_active=True)
        .exclude(batch_token=current_batch_token)
        .count()
    )


def _dry_run_apply(
    aggregated: dict[tuple[str, str, str, str, str, Decimal | None, int | None, str], ParsedStockRow],
    qty_by_key: dict[tuple[str, str, str, str, str, Decimal | None, int | None, str], int],
    product_stats: dict[tuple[str, str, str, str], dict[str, Any]],
    report: StockImportReport,
) -> None:
    series_seen: set[str] = set()
    category_seen: set[str] = set()
    city_seen: set[str] = set()
    product_seen: set[str] = set()

    for key, sample_row in aggregated.items():
        brand_slug, category_slug, model_code, config, city_slug, price, year, vat = key
        series_slug = (brand_slug or "other").lower() or "other"
        if series_slug not in series_seen:
            series_seen.add(series_slug)
            if not Series.objects.filter(slug__iexact=series_slug).exists():
                report.created_series += 1

        category_slug = (category_slug or "tehnika").lower() or "tehnika"
        if category_slug not in category_seen:
            category_seen.add(category_slug)
            if not Category.objects.filter(slug__iexact=category_slug).exists():
                report.created_categories += 1

        city_slug = (city_slug or "").lower() or _normalize_city_slug(sample_row.city_name)
        if city_slug not in city_seen:
            city_seen.add(city_slug)
            if not City.objects.filter(slug__iexact=city_slug).exists():
                report.created_cities += 1

        # Product identity
        sku, _slug = _product_identity(
            brand_slug=series_slug,
            category_slug=category_slug,
            model_code=model_code,
            config=config,
        )
        sku_ci = sku.lower()
        if sku_ci not in product_seen:
            product_seen.add(sku_ci)
            if Product.objects.filter(sku__iexact=sku).exists():
                report.updated_products += 1
            else:
                report.created_products += 1

        # Offer identity: depends on product+city
        # We can only approximate counts in dry-run without creating the product.
        # Use DB existence checks by joining on product.sku.
        total_qty = int(qty_by_key[key])
        _ = total_qty
        stats = product_stats.get((brand_slug, category_slug, model_code, config)) or {}
        _ = stats

        existing_product = Product.objects.filter(sku__iexact=sku).first()
        if existing_product:
            existing_city = City.objects.filter(slug__iexact=city_slug).first()
            if existing_city:
                if price is None:
                    exists = Offer.objects.filter(
                        product=existing_product,
                        city=existing_city,
                        year=year,
                        vat=vat,
                        price__isnull=True,
                    ).exists()
                else:
                    exists = Offer.objects.filter(
                        product=existing_product,
                        city=existing_city,
                        price=price,
                        year=year,
                        vat=vat,
                    ).exists()
                if exists:
                    report.updated_offers += 1
                else:
                    report.created_offers += 1
            else:
                # city will be created, offer too
                report.created_offers += 1
        else:
            report.created_offers += 1


def _load_workbook(file: str | Path | BinaryIO, import_logger: logging.Logger) -> openpyxl.Workbook:
    try:
        if hasattr(file, "seek"):
            try:
                file.seek(0)
            except Exception:  # noqa: BLE001
                pass
        return openpyxl.load_workbook(file)
    except Exception as exc:  # noqa: BLE001
        import_logger.exception("Failed to open workbook: %s", exc)
        raise


def _select_sheet(workbook: openpyxl.Workbook, sheet: str | None, import_logger: logging.Logger):
    if not sheet:
        return workbook.active
    if sheet in workbook.sheetnames:
        return workbook[sheet]
    import_logger.error("Sheet '%s' not found. Available: %s", sheet, workbook.sheetnames)
    raise ValueError(f"Sheet '{sheet}' not found")


def _get_import_logger() -> logging.Logger:
    logger = logging.getLogger(IMPORT_LOGGER_NAME)
    logger.setLevel(logging.INFO)
    return logger


_RU_TRANSLIT: dict[str, str] = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "yo",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "kh",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "shch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}


def _transliterate_ru(text: str) -> str:
    out: list[str] = []
    for ch in str(text):
        lower = ch.lower()
        if lower in _RU_TRANSLIT:
            out.append(_RU_TRANSLIT[lower])
        else:
            out.append(ch)
    return "".join(out)


def _slugify_any(text: str) -> str:
    """Slugify for ASCII slugs, with RU transliteration fallback."""

    base = slugify(text)
    if base:
        return base
    translit = _transliterate_ru(text)
    return slugify(translit)
