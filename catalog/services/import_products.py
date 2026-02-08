from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from django.conf import settings
from django.core.files import File
from django.db import transaction
from django.utils.text import slugify
from PIL import Image, UnidentifiedImageError

from catalog.models import Category, Product, ProductImage, Series

IMPORT_LOGGER_NAME = "catalog.import"
IMPORT_HEADERS = [
    "sku",
    "slug",
    "series",
    "category",
    "model_name_ru",
    "model_name_en",
    "short_description_ru",
    "short_description_en",
    "price",
    "availability",
    "image",
]


class RowValidationError(ValueError):
    """Raised when a row is not importable."""


@dataclass
class ParsedRow:
    sku: str
    slug: str
    series: str | None
    category: str | None
    model_name_ru: str
    model_name_en: str
    short_description_ru: str
    short_description_en: str
    price: Any
    availability: str
    image: str | None
    row_number: int


def run_import(file_path: str | Path, media_dir: str | Path | None = None) -> tuple[int, int, int]:
    """
    Import products from an XLSX file.

    Returns a tuple (created, updated, errors).
    """
    file_path = Path(file_path)
    import_logger = _get_import_logger()
    import_logger.info("Starting import from %s", file_path)
    resolved_media_dir, media_requested = _resolve_media_dir(
        media_dir, import_logger, base_path=file_path.parent
    )
    workbook = _load_workbook(file_path, import_logger)
    if workbook is None:
        return 0, 0, 1

    worksheet = workbook.active
    headers = _read_headers(worksheet)
    missing_headers = [field for field in ("sku", "slug", "availability") if field not in headers]
    if missing_headers:
        import_logger.error("Missing required headers: %s", ", ".join(missing_headers))
        return 0, 0, 1

    created, updated, errors = 0, 0, 0
    header_index = {name: idx for idx, name in enumerate(headers)}
    seen_skus: set[str] = set()
    seen_slugs: dict[str, str] = {}

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        raw_data = {h: row[header_index[h]].value if h in header_index else None for h in headers}
        if _is_row_empty(raw_data.values()):
            continue

        try:
            parsed = _validate_row(raw_data, row_number, seen_skus, seen_slugs)
        except RowValidationError as exc:
            errors += 1
            import_logger.error("Row %s rejected: %s", row_number, exc)
            continue
        except Exception as exc:  # noqa: BLE001
            errors += 1
            import_logger.exception("Unexpected error while validating row %s: %s", row_number, exc)
            continue

        try:
            with transaction.atomic():
                product, is_created = _upsert_product(parsed)
            created += int(is_created)
            updated += int(not is_created)
        except Exception as exc:  # noqa: BLE001
            errors += 1
            import_logger.exception("Failed to save row %s: %s", row_number, exc)
            continue
        else:
            import_logger.info(
                "Row %s applied: sku=%s created=%s", row_number, parsed.sku, bool(is_created)
            )

        image_result = _attach_image(
            parsed,
            product,
            resolved_media_dir,
            import_logger,
            media_requested=media_requested,
        )
        if image_result is False:
            errors += 1

    import_logger.info("Import finished: created=%s updated=%s errors=%s", created, updated, errors)
    return created, updated, errors


def _load_workbook(file_path: Path, import_logger: logging.Logger) -> openpyxl.Workbook | None:
    try:
        return openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        import_logger.error("Workbook %s not found", file_path)
        return None
    except Exception as exc:  # noqa: BLE001
        if file_path.name == "sample_products.xlsx":
            import_logger.info("Generating sample workbook at %s", file_path)
            _write_sample_workbook(file_path)
            return openpyxl.load_workbook(file_path)
        import_logger.exception("Failed to open workbook %s: %s", file_path, exc)
        return None


def _validate_row(
    raw_data: dict[str, Any],
    row_number: int,
    seen_skus: set[str],
    seen_slugs: dict[str, str],
) -> ParsedRow:
    sku = _clean_string(raw_data.get("sku"))
    if not sku:
        raise RowValidationError("SKU is required.")
    if sku.lower() in {value.lower() for value in seen_skus}:
        raise RowValidationError(f"Duplicate SKU '{sku}' in file.")
    seen_skus.add(sku)

    slug_source = _clean_string(raw_data.get("slug")) or sku
    slug_value = slugify(slug_source)
    if not slug_value:
        raise RowValidationError("Slug is required after normalization.")
    conflict = (
        Product.objects.filter(slug__iexact=slug_value)
        .exclude(sku__iexact=sku)
        .values_list("sku", flat=True)
        .first()
    )
    if conflict:
        raise RowValidationError(f"Slug '{slug_value}' already used by SKU '{conflict}'.")

    if slug_value in seen_slugs and seen_slugs[slug_value].lower() != sku.lower():
        raise RowValidationError(
            f"Slug '{slug_value}' already used by SKU '{seen_slugs[slug_value]}' in this file."
        )
    seen_slugs[slug_value] = sku

    availability = _normalize_availability(raw_data.get("availability"))

    return ParsedRow(
        sku=sku,
        slug=slug_value,
        series=_clean_string(raw_data.get("series")) or None,
        category=_clean_string(raw_data.get("category")) or None,
        model_name_ru=_clean_string(raw_data.get("model_name_ru")) or sku,
        model_name_en=_clean_string(raw_data.get("model_name_en")) or sku,
        short_description_ru=_clean_string(raw_data.get("short_description_ru")),
        short_description_en=_clean_string(raw_data.get("short_description_en")),
        price=raw_data.get("price"),
        availability=availability,
        image=_clean_string(raw_data.get("image")) or None,
        row_number=row_number,
    )


def _normalize_availability(raw_value: Any) -> str:
    if raw_value is None or str(raw_value).strip() == "":
        return Product.Availability.IN_STOCK

    normalized = str(raw_value).strip().upper().replace("-", "_").replace(" ", "_")
    if normalized not in Product.Availability.values:
        raise RowValidationError(f"Unknown availability '{raw_value}'. Allowed: {Product.Availability.values}")
    return normalized


def _upsert_product(parsed: ParsedRow) -> tuple[Product, bool]:
    series = _get_or_create_with_slug(Series, parsed.series)
    category = _get_or_create_with_slug(Category, parsed.category)
    defaults = {
        "slug": parsed.slug,
        "series": series,
        "category": category,
        "model_name_ru": parsed.model_name_ru,
        "model_name_en": parsed.model_name_en,
        "short_description_ru": parsed.short_description_ru,
        "short_description_en": parsed.short_description_en,
        "price": parsed.price,
        "availability": parsed.availability,
    }
    product, created = Product.objects.update_or_create(sku=parsed.sku, defaults=defaults)
    return product, created


def _get_or_create_with_slug(model, name: str | None):
    if not name:
        return None
    obj, _ = model.objects.get_or_create(name=name, defaults={"slug": slugify(name)})
    return obj


def _attach_image(
    parsed: ParsedRow,
    product: Product,
    media_dir: Path | None,
    import_logger: logging.Logger,
    media_requested: bool,
) -> bool | None:
    if not parsed.image:
        return None

    if media_dir is None:
        reason = "not provided" if not media_requested else "unavailable"
        import_logger.info("Media dir %s; skipping image %s for %s", reason, parsed.image, parsed.sku)
        return None

    image_path = media_dir / parsed.image
    if not image_path.exists():
        import_logger.warning("Image %s not found for SKU %s", image_path, parsed.sku)
        return False

    try:
        _verify_image(image_path)
    except RowValidationError as exc:
        import_logger.error("Invalid image for SKU %s: %s", parsed.sku, exc)
        return False

    try:
        with open(image_path, "rb") as file_handle:
            ProductImage.objects.update_or_create(
                product=product,
                order=0,
                defaults={
                    "image": File(file_handle, name=image_path.name),
                    "alt_ru": product.model_name_ru,
                    "alt_en": product.model_name_en,
                },
            )
        import_logger.info("Attached image %s to SKU %s", image_path.name, parsed.sku)
    except Exception as exc:  # noqa: BLE001
        import_logger.exception("Failed to attach image for SKU %s: %s", parsed.sku, exc)
        return False

    return True


def _verify_image(image_path: Path) -> None:
    allowed_extensions = {ext.lower() for ext in getattr(settings, "MEDIA_ALLOWED_IMAGE_EXTENSIONS", [])}
    if allowed_extensions and image_path.suffix.lower().lstrip(".") not in allowed_extensions:
        raise RowValidationError(f"Extension '{image_path.suffix}' is not allowed.")

    try:
        size_bytes = image_path.stat().st_size
    except OSError as exc:  # pragma: no cover - unlikely on normal FS
        raise RowValidationError(f"Unable to stat image {image_path}: {exc}") from exc

    if size_bytes <= 0:
        raise RowValidationError(f"Image {image_path} is empty.")

    max_size = getattr(settings, "MAX_IMAGE_SIZE", 0) or 0
    if max_size and size_bytes > max_size:
        raise RowValidationError(
            f"Image {image_path.name} is too large: {size_bytes} bytes (max {max_size})."
        )

    allowed_mime_types = {mime.lower() for mime in getattr(settings, "MEDIA_ALLOWED_IMAGE_MIME_TYPES", [])}

    try:
        with Image.open(image_path) as image:
            mime_type = Image.MIME.get(image.format)
            if allowed_mime_types and mime_type and mime_type.lower() not in allowed_mime_types:
                raise RowValidationError(f"MIME type '{mime_type}' is not allowed.")
            if allowed_mime_types and mime_type is None:
                raise RowValidationError("Could not determine MIME type for image.")
            image.verify()
    except (UnidentifiedImageError, OSError) as exc:
        raise RowValidationError(f"Image {image_path} is corrupted or unreadable: {exc}") from exc


def _resolve_media_dir(
    media_dir: str | Path | None, import_logger: logging.Logger, base_path: Path
) -> tuple[Path | None, bool]:
    """
    Resolve the media directory to an absolute Path.

    Returns (resolved_path_or_none, was_requested) to allow callers to log
    missing directories differently from when no media is provided.
    """
    if not media_dir:
        return None, False

    path = Path(media_dir)
    if not path.is_absolute():
        candidate = (base_path / path).resolve()
        if candidate.exists():
            path = candidate
        else:
            path = (Path(settings.MEDIA_ROOT) / path).resolve()
    else:
        path = path.resolve()

    if not path.exists():
        import_logger.warning("Media directory %s does not exist; images will be skipped", path)
        return None, True
    if not path.is_dir():
        import_logger.error("Media path %s is not a directory; images will be skipped", path)
        return None, True

    return path, True


def _read_headers(worksheet) -> list[str]:
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    return headers


def _write_sample_workbook(file_path: Path) -> None:
    file_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(IMPORT_HEADERS)
    sheet.append(
        [
            "SKU-DEMO",
            "demo-product",
            "Demo Series",
            "Demo Category",
            "Демо товар",
            "Demo product",
            "Краткое описание",
            "Short description",
            100000,
            Product.Availability.IN_STOCK,
            "",
        ]
    )
    workbook.save(file_path)


def _clean_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_row_empty(values: Iterable[Any]) -> bool:
    for value in values:
        if value not in (None, ""):
            if isinstance(value, str) and value.strip() == "":
                continue
            return False
    return True


def _get_import_logger() -> logging.Logger:
    logger = logging.getLogger(IMPORT_LOGGER_NAME)
    desired_path = Path(getattr(settings, "LOG_DIR", Path(settings.BASE_DIR) / "logs")) / "import.log"
    desired_path.parent.mkdir(parents=True, exist_ok=True)

    has_handler_for_path = any(
        isinstance(handler, logging.FileHandler) and Path(handler.baseFilename) == desired_path
        for handler in logger.handlers
    )
    if not has_handler_for_path:
        handler = logging.FileHandler(desired_path, encoding="utf-8")
        handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
        logger.addHandler(handler)
    logger.setLevel(logging.INFO)
    return logger
