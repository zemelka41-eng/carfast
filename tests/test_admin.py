from io import BytesIO

import openpyxl
import pytest
from django.contrib.admin.sites import AdminSite

from catalog.admin import ProductAdmin
from catalog.models import Product

pytestmark = pytest.mark.django_db


@pytest.fixture
def product_admin():
    return ProductAdmin(Product, AdminSite())


def test_make_published_and_unpublished_actions(product_admin, admin_request, product):
    queryset = Product.objects.filter(pk=product.pk)

    product_admin.make_published(admin_request, queryset)
    product.refresh_from_db()
    assert product.published is True

    product_admin.make_unpublished(admin_request, queryset)
    product.refresh_from_db()
    assert product.published is False


def test_export_xlsx_returns_workbook(product_admin, admin_request, product):
    queryset = Product.objects.filter(pk=product.pk)

    response = product_admin.export_xlsx(admin_request, queryset)
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = openpyxl.load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 2  # header + one row
    values = [cell.value for cell in ws[2]]
    assert product.sku in values
    assert product.model_name_ru in values


def test_import_sample_calls_importer(monkeypatch, product_admin, admin_request):
    calls = {}

    def fake_run_import(file_path, media_dir):
        calls["args"] = (file_path, media_dir)
        return 1, 0, 0

    monkeypatch.setattr("catalog.admin.importers.run_import", fake_run_import)

    product_admin.import_sample(admin_request, Product.objects.none())

    assert "args" in calls
    assert list(admin_request._messages)
